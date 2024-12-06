import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def LevantarNumeros(ruta):
    # Ruta del archivo (puede ser un CSV o un Excel)
    with pd.ExcelFile(ruta) as xls:
    # Leer el archivo (ajusta el método si es Excel)
        df = pd.read_excel(xls)  # Usa pd.read_excel() si es un archivo .xlsx
    try:
        return df["Telefono"]
    except:
        return False


def LevantarNombres(ruta):
    # Ruta del archivo (puede ser un CSV o un Excel)
    ruta_archivo = ruta  # Cambia esto por el nombre de tu archivo

    # Leer el archivo (ajusta el método si es Excel)
    df = pd.read_excel(ruta_archivo)  # Usa pd.read_excel() si es un archivo .xlsx
    try:
        return df["Nombre"]
    except:
        return False

#pintar numeros ya usados
def PintarUsados(ruta, nombre_columna, numero_telefono):
    color = "FFFF00"  # color amarillo
    # Cargar el archivo y la hoja activa
    wb = load_workbook(ruta)
    ws = wb.active  # Utiliza la hoja activa del archivo
    numero_telefono = numero_telefono.replace("+", "")
    # Obtener la columna por nombre
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == nombre_columna:
            columna = col[0].column
            letra_columna = ws.cell(row=1, column=columna).column_letter
            break

    # Buscar el número de teléfono en la columna
    for fila in range(2, ws.max_row + 1):
        celda = ws[f"{letra_columna}{fila}"]
        if str(celda.value).strip() == str(numero_telefono).strip():
            # Pintar la celda
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            celda.fill = fill
            break

    # Guardar el archivo
    wb.save(ruta)


def CeldaPintada(ruta, nombre_columna):
    color = "FFFF00" # color que va a buscar
    #archivo que va a trabajar
    wb = load_workbook(ruta)
    #lo activa
    ws= wb.active

    columna = None
    #itera sobre las columnas del archivo
    for col in ws.iter_cols(1, ws.max_column):
        #busca el nombre del archivo
        if col[0].value == nombre_columna:
            columna = col[0].column
            break
    # guarda la letra de la columna cn el nombre
    letra_columna = ws.cell(row=1,column=columna).column_letter

    numeros_no_pintados = []
    hay_pintados = False
    
    #recorre la fila seleccionada
    for fila in range(2,ws.max_row+1):
        #define la celda sumando la letra con en N°
        celda = f"{letra_columna}{fila}"
        #selecciona la celda
        celda_actual=ws[celda]
        #pregunta si esta pintado de amarillo
        pintada = (
            celda_actual.fill.start_color.rgb == f"FF{color}"  # Verifica el color de fondo
            if celda_actual.fill.start_color and celda_actual.fill.fill_type == "solid"
            else False
        )
        # establece si hay o no pintadas         
        if pintada:
            hay_pintados = True  # Se encontró al menos una celda pintada

        # Si la celda no está pintada, agregar el número de la columna "Telefono" a la lista
        if not pintada:
            telefono = ws[celda].value
            numeros_no_pintados.append(telefono)

    #retorna una lista de dos valores, un bool y una lista
    return hay_pintados, numeros_no_pintados

if __name__ == "__main__":
    PintarUsados("C:\\Users\\Sistemas\\Desktop\\AutoSend\\Files\\test.xlsx", "Telefono", "543385448553")
    #print(CeldaPintada(ruta= "C:\\Users\\Sistemas\\Desktop\\AutoSend\\Files\\test.xlsx",
    #             nombre_columna= "Telefono"))
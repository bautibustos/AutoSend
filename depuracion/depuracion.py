import os
import sys
import webbrowser
import time
import subprocess
import openpyxl
import threading
import json
from pywinauto import Desktop
from collections import Counter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import modulos.utils as utils
import modulos.ManejoExel as LE  # Ahora debe encontrarlo correctamente


#path raiz
path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))+"\\files"


def DepuracionCompleta(file):
    with open(f"{os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))}\\config.json", 'r', encoding='utf-8') as archivo:
        tiempo = json.load(archivo)["tiempo_depuracion"]

    path_file = f"{path}\\{file}"

    # Preguntamos si está la columna "Telefono"|
    if LE.LevantarNumeros(path_file) is False:
        utils.ErrorCarga(error_mensaje='Columna "Telefono" No encontrada')    
    else:
        # Obtener los números de teléfono
        numeros_de_telefono = LE.LevantarNumeros(path_file)

        #comprovacion de la existecia de la columna nombre
        if LE.LevantarNombres(path_file) is False:
            utils.ErrorCarga(error_mensaje='Columna "Nombre" No encontrada')
        else:
            #levanamos los nombres 
            nombres = LE.LevantarNombres(path_file)

        ###################### filtrado que sean superiores a 10 ######################
        numeros_filtrados = []
        nombres_validos = []
        contador = 0
        erroneos = []
        #recorremos y guardamos los numeros mayores a 10 digitos
        for numero in numeros_de_telefono:
            if len(str(numero)) >= 10 and str(numero) != "nan": 
                            
                item = str(numero)#lo convierto enstr
                if not "+54" in item:#me fijo tiene el mas +54
                    #comparo si los primeros dos items son 54
                    if not "54" in item[:2]:
                        item = "54"+item#agro el 54 si no lo tiene
                    if not "+" in item:# pregunto si tengo el +
                        item = "+"+item# agrego el + si no esta
                    # Elimino guiones de por medio
                item = item.replace('-', '')
                
                # Elimino espacios si hay
                item = item.replace(' ', '')
                
                # Limpiar paréntesis
                item = item.replace('(', '')
                
                item = item.replace(')', '')
            
                #sumamos a los validos en listas diferentes
                numeros_filtrados.append(str(item))
                nombres_validos.append(nombres[contador])
            else:
                # sumamos a una lista los numeros erroneos con el nombre
                erroneos.append([numero, nombres[contador]])
            contador += 1

        # libreria para contar repeticiones
        contador = Counter(numeros_filtrados)
        # Inicializar listas
        no_duplicados = []
        duplicados = []

        # Recorrer la lista original
        for tel, nom in zip(numeros_filtrados, nombres_validos):
            if contador[tel] == 1:  # Si el número aparece solo una vez, guardarlo en no_duplicados
                no_duplicados.append([tel, nom])
            else:  # Si el número aparece más de una vez, guardarlo en duplicados
                duplicados.append([tel, nom])

        ################# validador whatsapp ######################
        # nombre de las columnas
        tiene = [["Telefono","Nombre"]]
        # es un return estilo flag probocada por un hilo
        error_event = threading.Event()
        # revision del error
        stop_deamon = threading.Event()
        def revisa_error():
            while not stop_deamon.is_set():
        # ...   utilizando ventanas lista todos los procesos de windows
                ventanas = Desktop(backend="win32").windows()
                #recorro la lista en busca de una ventana
                for ventana in ventanas:
                    try:
                        # obitene el nombre de la clase
                        clase = ventana.class_name()
                        # Filtrar por clase o título.
                        #       la clase xalm windowedpopupclass es un proceso de windows de popups al momento de no tener whatsapp 
                        if clase == "Xaml_WindowedPopupClass":
                            # Finalizar un proceso por su nombre, en este caso finalizando whatsapp el proceso de popup desaparece
                            subprocess.run(
                                ["taskkill", "/F", "/IM", "whatsapp.exe"], 
                                check=True, 
                                stdout=subprocess.DEVNULL, 
                                stderr=subprocess.DEVNULL
                            ) #stdout y stderr poniendlos en devnull evitan que retorne el mensaje si lo cerro o no

                            #marcamos la bandera
                            error_event.set()
                    except Exception as e:
                        pass
                time.sleep(1)

        # se ejecuta el hilo que revisa continuamente el proceso de whastapp
        validador = threading.Thread(target=revisa_error, daemon=True)
        validador.start()
        
        # recorre y abre los numeros de whatsapp
        for numero in no_duplicados:
            webbrowser.open(f'whatsapp://send?phone={str(numero[0])}')
            time.sleep(tiempo)
            #si la bandera esta activa es porque el numero no es valido
            if error_event.is_set():
                #resetea la bandera
                error_event.clear()
                #agrega al registro el numero erroneo
                erroneos.append(numero)
            else:
                #si el numero es valido se guarda
                tiene.append(numero)
        stop_deamon.set()
        validador.join()

        # funcion para guardar los exels 
        def exel(lista, nombre):
            #abrimos el archivo mod, es donde se guardan los numeros y nombres filtrados
            wb = openpyxl.load_workbook(os.getcwd()+"\\depuracion\\dep.xlsx")
            ws = wb.active
            #guaradmos en el exel
            for i in range(0, len(lista)):
                ws[f'A{i+1}']=(lista[i][0])
                ws[f'B{i+1}']=(lista[i][1])
            
            wb.save(path+f"\\Depurado - {nombre}.xlsx")
            wb.close()
        
        for i in erroneos:
            duplicados.append(i)

        exel(duplicados, nombre="Erroneos")
        exel(tiene, nombre="Validos")
    
if __name__ == "__main__":
    DepuracionCompleta("rio_tercero_p2.xlsx")


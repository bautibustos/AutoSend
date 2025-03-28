import os
import sys

# Agregar la carpeta raíz del proyecto al sys.path, esto tambien ayuda a que pueda levantar modulos que estan en la careta raiz
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import modulos.ManejoExel as LE  # Ahora debe encontrarlo correctamente
import openpyxl

# se obtiene la ruta raiz del proyecto, y luego se posiciona en la carpeta files
path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))+"\\files"


def DepuracionNumeros(nombre_archivo):
    # armamos la ubicacion con el nombre del archivo
    path_file = f"{path}\\{nombre_archivo}"
    #guardamos los numeros de telefono y los nombres en listas
    Cantidad_Numero = LE.LevantarNumeros(path_file)
    Nombres = LE.LevantarNombres(path_file)

    Validos = []
    NombreValido = []
    contador = 0
    #recorremos y guardamos los numeros mayores a 10 digitos
    for numero in Cantidad_Numero:
        if len(str(numero)) >= 10:
            Validos.append(numero)
            NombreValido.append(Nombres[contador])
        contador += 1

    #abrimos el archivo mod, es donde se guardan los numeros y nombres filtrados
    openpyxl.load_workbook(path+"\\Mod.xlsx")
    wb = openpyxl.load_workbook( path+"\\Mod.xlsx")
    ws = wb.active
    #guaradmos en el exel
    for i in range(len(Validos)):
        ws[f'A{i+1}']=(NombreValido[i])
        ws[f'B{i+1}']=(Validos[i])
    
    wb.save(path+"\\Mod.xlsx")
    wb.close()

if __name__ == "__main__":
    DepuracionNumeros("comunitarios-filtrado.xlsx")